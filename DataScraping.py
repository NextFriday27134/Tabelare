# De adaugat: threading
# Pt chirie trebuie adaugat si linkul catre postare
#                 ¯\_(ツ)_/¯

from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.cell import get_column_letter


REAL_ESTATE_SITE_SALE = "https://www.storia.ro/ro/rezultate/vanzare/apartament/iasi?limit=72"
REAL_ESTATE_SITE_RENT = "https://www.storia.ro/ro/rezultate/inchiriere/apartament/iasi?limit=72"
SITE_HOME_PAGE = "https://www.storia.ro"
HEADERS_SALE = ["Zona", "Pret", 'Suprafață', 'Forma de proprietate', 'Numărul de camere', 'Stare', 'Etaj',
                'Balcon/grădină/terasă', 'Chirie', 'Garaj/loc de parcare', 'Vizionare la distanță', 'Încălzire',
                'Tip proprietate', 'Tip vânzător', 'Liber de la', 'Anul construcției', 'Tip clădire', 'Tip geamuri',
                'Lift', 'Media', 'Siguranță', 'Facilități', 'Informații suplimentare', 'Material de construcție',
                "Titlu", "Descriere"]
HEADERS_RENT = ["Zona", "Pret", 'Suprafață', 'Chirie', 'Numărul de camere', 'Depozit', 'Etaj', 'Tip clădire',
                'Liber de la', 'Balcon/grădină/terasă', 'Vizionare la distanță', 'Stare', 'Tip vânzător',
                'Acceptă studenți', 'Facilități', 'Media', 'Încălzire', 'Siguranță', 'Tip geamuri', 'Lift',
                'Garaj/loc de parcare', 'Anul construcției', 'Material de construcție', 'Informații suplimentare',
                "Titlu", "Descriere"]
ALL_APARTMENTS_URLS = []


def get_site_html_code(site_url):
    response = requests.get(site_url)
    return response.text


def make_it_beauty(response_text):
    soup = BeautifulSoup(response_text, "html.parser")
    return soup


def get_new_soup(page_number, real_estate_site):
    new_url = real_estate_site + f'&page={page_number}'
    new_soup = make_it_beauty(get_site_html_code(new_url))
    return new_soup


def get_all_apartments_url_from_current_page(soup_object: BeautifulSoup):
    apartments_list = []
    posts_apartments_list = soup_object.find_all(name='a', class_="css-lsw81o")
    for apartment in posts_apartments_list:
        apartments_list.append(apartment.get("href"))
    return apartments_list


def get_all_apartments_url(soup_object: BeautifulSoup, real_estate_site):
    global ALL_APARTMENTS_URLS
    last_page_nr = int(soup_object.find_all(name='button', class_="css-j4ip63")[-2].getText())
    ALL_APARTMENTS_URLS.extend(get_all_apartments_url_from_current_page(soup_object))
    for page_nr in range(2, last_page_nr+1):
        new_soup = get_new_soup(page_nr, real_estate_site)
        ALL_APARTMENTS_URLS.extend(get_all_apartments_url_from_current_page(new_soup))


def get_data_from_apartment_page(soup_object: BeautifulSoup):
    data_from_page = soup_object.find_all(name='div', class_="enb64yk1")
    data_from_tables = [data.getText("#").split("#")[1] for data in data_from_page]
    apartment_price = soup_object.find(name='strong', class_="e1l1avn10", attrs={"aria-label": "Preț"})
    data_from_tables.insert(0, apartment_price.getText())
    apartment_zone = soup_object.find(name='a', class_="e1w8sadu0", attrs={'aria-label': "Abordare"})
    data_from_tables.insert(0, apartment_zone.getText())
    apartment_title = soup_object.find(name='h1', class_='efcnut38')
    data_from_tables.append(apartment_title.getText())
    apartment_description = soup_object.find(name='div', class_="e1lbnp621", attrs={"data-cy": "adPageAdDescription"})
    data_from_tables.append(apartment_description.getText())
    return data_from_tables


def save_data_as_excel_table(table_name, header):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for url in ALL_APARTMENTS_URLS:
        new_url = SITE_HOME_PAGE + url
        soup = make_it_beauty(get_site_html_code(new_url))
        try:
            ws.append(get_data_from_apartment_page(soup))
        except AttributeError:
            # pare o eroare de la server ar merge sa pun sa incerce de cateva ori din nou pt intrarile respective
            #                                  (-_-)
            print(new_url)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    tab = Table(displayName=table_name, ref=f"A1:{get_column_letter(len(HEADERS_SALE))}{len(ALL_APARTMENTS_URLS) + 1}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    wb.save(f"{table_name}.xlsx")


if __name__ == "__main__":
    # List of apartments for sale

    soup = make_it_beauty(get_site_html_code(REAL_ESTATE_SITE_SALE))
    get_all_apartments_url(soup, REAL_ESTATE_SITE_SALE)
    save_data_as_excel_table(table_name='apartamente_de_vanzare', header=HEADERS_SALE)

    # List of apartments for rent.

    soup = make_it_beauty(get_site_html_code(REAL_ESTATE_SITE_RENT))
    get_all_apartments_url(soup, REAL_ESTATE_SITE_RENT)
    save_data_as_excel_table(table_name='apartamente_de_inchiriat', header=HEADERS_RENT)
